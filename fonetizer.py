#!/usr/bin/env python3
"""
Fonetizer: Convert song lyrics to singing-phonetic tables

Usage:
    python fonetizer.py input.txt
    cat lyrics.txt | python fonetizer.py
"""

import sys
import re
import os
from typing import List, Tuple
import cmudict
import eng_to_ipa as ipa_converter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Load CMU Pronouncing Dictionary
CMU_DICT = cmudict.dict()

# ARPABET to IPA mapping (with stress-based vowel length)
ARPABET_TO_IPA = {
    # Vowels - primary stress (1) = long vowels
    'AA': 'ɑ',    'AA0': 'ɑ',    'AA1': 'ɑː',   'AA2': 'ɑ',
    'AE': 'æ',    'AE0': 'æ',    'AE1': 'æ',    'AE2': 'æ',
    'AH': 'ʌ',    'AH0': 'ə',    'AH1': 'ʌ',    'AH2': 'ʌ',
    'AO': 'ɔ',    'AO0': 'ɔ',    'AO1': 'ɔː',   'AO2': 'ɔ',
    'AW': 'aʊ',   'AW0': 'aʊ',   'AW1': 'aʊ',   'AW2': 'aʊ',
    'AY': 'aɪ',   'AY0': 'aɪ',   'AY1': 'aɪ',   'AY2': 'aɪ',
    'EH': 'ɛ',    'EH0': 'ɛ',    'EH1': 'ɛ',    'EH2': 'ɛ',
    'ER': 'ɜːr',  'ER0': 'ər',   'ER1': 'ɝː',   'ER2': 'ɜːr',
    'EY': 'eɪ',   'EY0': 'eɪ',   'EY1': 'eɪ',   'EY2': 'eɪ',
    'IH': 'ɪ',    'IH0': 'ɪ',    'IH1': 'ɪ',    'IH2': 'ɪ',
    'IY': 'i',    'IY0': 'i',    'IY1': 'iː',   'IY2': 'i',
    'OW': 'oʊ',   'OW0': 'oʊ',   'OW1': 'oʊ',   'OW2': 'oʊ',
    'OY': 'ɔɪ',   'OY0': 'ɔɪ',   'OY1': 'ɔɪ',   'OY2': 'ɔɪ',
    'UH': 'ʊ',    'UH0': 'ʊ',    'UH1': 'ʊ',    'UH2': 'ʊ',
    'UW': 'u',    'UW0': 'u',    'UW1': 'uː',   'UW2': 'u',

    # Consonants
    'B': 'b',   'CH': 'ʧ',  'D': 'd',   'DH': 'ð',
    'F': 'f',   'G': 'g',   'HH': 'h',  'JH': 'ʤ',
    'K': 'k',   'L': 'l',   'M': 'm',   'N': 'n',
    'NG': 'ŋ',  'P': 'p',   'R': 'r',   'S': 's',
    'SH': 'ʃ',  'T': 't',   'TH': 'θ',  'V': 'v',
    'W': 'w',   'Y': 'j',   'Z': 'z',   'ZH': 'ʒ',
}


def arpabet_to_ipa(arpabet_list: List[str]) -> str:
    """Convert ARPABET phonemes to IPA string"""
    ipa = ""
    for phoneme in arpabet_list:
        if phoneme in ARPABET_TO_IPA:
            ipa += ARPABET_TO_IPA[phoneme]
        else:
            # Try without stress marker
            base = ''.join([c for c in phoneme if not c.isdigit()])
            if base in ARPABET_TO_IPA:
                ipa += ARPABET_TO_IPA[base]
    return ipa

# Diphthongs that need to be split: [long vowel] + [glide]
DIPHTHONGS = {
    "aɪ": ("aː", "ɪ"),  # cry, I, my, night
    "eɪ": ("eː", "ɪ"),  # way, day, came, they
    "ɔɪ": ("ɔː", "ɪ"),  # boy, toy
    "aʊ": ("aː", "ʊ"),  # now, how
    "oʊ": ("oː", "ʊ"),  # know, old, tomorrow
    "ɪə": ("ɪː", "ə"),  # here, dear
    "ɛə": ("ɛː", "ə"),  # there, care
    "ʊə": ("ʊː", "ə"),  # sure, tour
}

# IPA vowels (expanded for CMU/eng_to_ipa output)
VOWELS = set("aeiouæɒɔəɜʌɪʊɛɑɝː")

# IPA consonants (expanded for CMU/eng_to_ipa output)
CONSONANTS = set("bcdfghjklmnpqrstvwxyzðθʃʒŋʧʤ")


def parse_input_line(line: str) -> Tuple[str, str]:
    """
    Parse input line format: N text or just text

    Where N is the starting measure number (optional, no T prefix, no end measure).

    Returns:
        (start_measure, text) - start_measure is empty string if not provided
    """
    match = re.match(r'(\d+)\s+(.+)', line.strip())
    if match:
        return match.group(1), match.group(2)

    # No measure number - just return the text with empty measure
    return "", line.strip()


def text_to_ipa(word: str) -> str:
    """
    Convert an English word to IPA notation.

    Priority:
    1. Manual singing-specific overrides
    2. CMU Pronouncing Dictionary (134k words, accurate)
    3. eng_to_ipa library (fallback for words not in CMU)
    """
    word_clean = word.lower().strip(".,!?;:'\"")

    if not word_clean:
        return ""

    # Manual overrides for singing pronunciation
    SINGING_OVERRIDES = {
        "used": "uːzd",  # Remove initial j for singing (juːzd → uːzd)
        "to": "tuː",     # Always long u: for singing
    }

    if word_clean in SINGING_OVERRIDES:
        return SINGING_OVERRIDES[word_clean]

    # Try CMU dictionary
    if word_clean in CMU_DICT:
        # Take first pronunciation (most common)
        arpabet = CMU_DICT[word_clean][0]
        ipa = arpabet_to_ipa(arpabet)
        return ipa

    # Fallback to eng_to_ipa
    try:
        ipa = ipa_converter.convert(word_clean)
        # Remove asterisks that indicate uncertain pronunciations
        ipa = ipa.replace("*", "")
        return ipa
    except Exception:
        # Last resort: return the word itself
        return word_clean


def find_vowel_positions(ipa: str) -> List[Tuple[int, int, str]]:
    """
    Find all vowels in IPA string and return their positions.
    Recognizes diphthongs as single vowel units.

    Returns:
        List of (start_pos, end_pos, vowel_string) tuples
    """
    vowels = []
    i = 0
    while i < len(ipa):
        if ipa[i] in VOWELS:
            start = i
            vowel = ipa[i]
            i += 1

            # Check for diphthong (two-character vowel pattern)
            if i < len(ipa) and ipa[i] in VOWELS:
                # Check if this forms a known diphthong
                potential_diphthong = vowel + ipa[i]
                if potential_diphthong in DIPHTHONGS:
                    vowel = potential_diphthong
                    i += 1

            # Include length marker if present
            if i < len(ipa) and ipa[i] == "ː":
                vowel += ipa[i]
                i += 1

            vowels.append((start, i, vowel))
        else:
            i += 1
    return vowels


def syllabify_phrase_ipa(phrase_ipa: str, is_phrase_final: bool) -> List[Tuple[str, str]]:
    """
    Syllabify a complete phrase IPA string.

    Algorithm:
    1. Find all vowels in the phrase
    2. For each vowel, assign consonants:
       - Consonants BEFORE vowel = this syllable's initial consonants
       - Consonants AFTER vowel = next syllable's initial consonants
       - Exception: phrase-final consonants get their own syllable with empty vowel
    3. Handle diphthongs by splitting into two rows

    Returns:
        List of (consonant, vowel) tuples
    """
    if not phrase_ipa:
        return []

    # Find all vowel positions
    vowel_positions = find_vowel_positions(phrase_ipa)

    if not vowel_positions:
        return []

    syllables = []

    # Process each vowel
    for vowel_idx, (vowel_start, vowel_end, vowel) in enumerate(vowel_positions):
        # Determine where consonants for this syllable start
        if vowel_idx == 0:
            cons_start = 0
        else:
            # Start after previous vowel
            _, prev_vowel_end, _ = vowel_positions[vowel_idx - 1]
            cons_start = prev_vowel_end

        # Collect consonants before this vowel
        consonants = ""
        for i in range(cons_start, vowel_start):
            if phrase_ipa[i] in CONSONANTS:
                consonants += phrase_ipa[i]

        # Handle diphthongs
        if vowel in DIPHTHONGS:
            long_v, glide = DIPHTHONGS[vowel]
            syllables.append((consonants, long_v))
            syllables.append(("", glide))
        else:
            syllables.append((consonants, vowel))

    # Handle phrase-final consonants
    if is_phrase_final and vowel_positions:
        _, last_vowel_end, _ = vowel_positions[-1]

        # Collect any remaining consonants
        final_consonants = ""
        for i in range(last_vowel_end, len(phrase_ipa)):
            if phrase_ipa[i] in CONSONANTS:
                final_consonants += phrase_ipa[i]

        # Add as separate syllable with empty vowel
        if final_consonants:
            syllables.append((final_consonants, ""))

    return syllables


def move_medial_consonants(syllables: List[Tuple[str, str]], is_phrase_final: bool) -> List[Tuple[str, str]]:
    """
    Move non-phrase-final consonants to the next syllable.
    Phrase-final consonants stay with their syllable.
    """
    if len(syllables) <= 1:
        return syllables

    result = []

    for i in range(len(syllables)):
        cons, vowel = syllables[i]

        # If this is the last syllable and phrase-final, keep consonants
        if i == len(syllables) - 1 and is_phrase_final:
            result.append((cons, vowel))
        # If there's a next syllable, check if we should move consonants
        elif i < len(syllables) - 1:
            next_cons, next_vowel = syllables[i + 1]

            # If current syllable has consonants after the vowel (shouldn't in our model)
            # or if next syllable has no consonant, this is handled by the model
            result.append((cons, vowel))
        else:
            result.append((cons, vowel))

    return result


def text_to_phonetic_syllables(text: str) -> List[Tuple[str, str]]:
    """
    Convert text to phonetic syllables as (consonant_cluster, vowel) pairs.

    Each tuple represents one row in the output table.
    Diphthongs are split into two tuples (second has empty consonant).

    Algorithm:
    1. Convert entire phrase to IPA (word by word, then concatenate)
    2. Syllabify the complete phrase IPA as one string
    3. This allows consonants to move across word boundaries

    Returns:
        List of (consonant, vowel) tuples
    """
    # Check if phrase-final
    text_clean = text.strip()
    is_phrase_final = text_clean.endswith(('.', '!', '?', ','))
    words = text_clean.split()

    # Convert all words to IPA and concatenate into single phrase IPA
    phrase_ipa = ""
    for word in words:
        word_ipa = text_to_ipa(word)
        phrase_ipa += word_ipa

    # Syllabify the complete phrase
    syllables = syllabify_phrase_ipa(phrase_ipa, is_phrase_final)

    return syllables


def process_phrase(text: str) -> List[Tuple[str, str]]:
    """
    Process a complete phrase into phonetic syllables.

    Returns:
        List of (consonant, vowel) tuples, one per table row
    """
    return text_to_phonetic_syllables(text)


def build_table(phrases: List[Tuple[str, List[Tuple[str, str]]]]) -> str:
    """
    Build the transposed table with column pairs for each phrase.

    Args:
        phrases: List of (start_measure, syllables)

    Returns:
        CSV formatted string
    """
    # Find maximum number of syllables (rows needed)
    max_syllables = max(len(syllables) for _, syllables in phrases)

    # Build header row
    header = ["Syllable"]
    for start, _ in phrases:
        header.extend([f"T{start}", ""])  # Only start measure on left column

    # Build data rows
    rows = [header]

    for syllable_idx in range(max_syllables):
        row = [str(syllable_idx + 1)]

        for _, syllables in phrases:
            if syllable_idx < len(syllables):
                consonant, vowel = syllables[syllable_idx]
                row.append(consonant)
                row.append(vowel)
            else:
                row.append("")  # Empty consonant
                row.append("")  # Empty vowel

        rows.append(row)

    # Convert to CSV
    return '\n'.join('\t'.join(row) for row in rows)


def build_excel_table(phrases: List[Tuple[str, List[Tuple[str, str]]]], output_path: str):
    """
    Build a formatted Excel table with column pairs for each phrase.

    Args:
        phrases: List of (start_measure, syllables)
        output_path: Path to save the Excel file

    Formatting:
        - Consonant columns: right-aligned
        - Vowel columns: left-aligned, bold
        - Borders around each column pair
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Phonetic Table"

    # Find maximum number of syllables (rows needed)
    max_syllables = max(len(syllables) for _, syllables in phrases) if phrases else 0

    # Border styles
    thin_border = Side(style='thin', color='000000')
    thick_border = Side(style='medium', color='000000')

    # Build header row
    header = ["Syllable"]
    for start, _ in phrases:
        header.extend([f"T{start}", ""])  # Only start measure on left column

    ws.append(header)

    # Format header row
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add top and bottom borders
        if col_idx == 1:
            # First column (Syllable)
            cell.border = Border(
                top=thick_border,
                bottom=thick_border,
                left=thick_border,
                right=thin_border
            )
        else:
            # Phrase columns - add thick border between pairs
            phrase_idx = (col_idx - 2) // 2
            is_first_in_pair = (col_idx - 2) % 2 == 0
            is_last_in_pair = (col_idx - 2) % 2 == 1
            is_last_column = col_idx == len(header)

            left = thick_border if is_first_in_pair else thin_border
            right = thick_border if (is_last_in_pair or is_last_column) else thin_border

            cell.border = Border(
                top=thick_border,
                bottom=thick_border,
                left=left,
                right=right
            )

    # Build data rows
    for syllable_idx in range(max_syllables):
        row_data = [str(syllable_idx + 1)]

        for _, syllables in phrases:
            if syllable_idx < len(syllables):
                consonant, vowel = syllables[syllable_idx]
                row_data.append(consonant)
                row_data.append(vowel)
            else:
                row_data.append("")  # Empty consonant
                row_data.append("")  # Empty vowel

        ws.append(row_data)

        # Format data row
        row_num = syllable_idx + 2  # +2 because header is row 1, and we're 0-indexed
        for col_idx, cell in enumerate(ws[row_num], start=1):
            is_last_row = syllable_idx == max_syllables - 1

            if col_idx == 1:
                # Syllable number column
                cell.alignment = Alignment(horizontal='center', vertical='center')
                bottom = thick_border if is_last_row else thin_border
                cell.border = Border(
                    top=thin_border,
                    bottom=bottom,
                    left=thick_border,
                    right=thin_border
                )
            else:
                # Phrase columns
                phrase_idx = (col_idx - 2) // 2
                is_consonant = (col_idx - 2) % 2 == 0
                is_vowel = (col_idx - 2) % 2 == 1
                is_first_in_pair = is_consonant
                is_last_in_pair = is_vowel
                is_last_column = col_idx == len(row_data)

                # Alignment
                if is_consonant:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:  # vowel
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = Font(bold=True)  # Bold vowels

                # Borders
                left = thick_border if is_first_in_pair else thin_border
                right = thick_border if (is_last_in_pair or is_last_column) else thin_border
                bottom = thick_border if is_last_row else thin_border

                cell.border = Border(
                    top=thin_border,
                    bottom=bottom,
                    left=left,
                    right=right
                )

    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_idx).column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max(max_length + 2, 8)  # Minimum width of 8
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_path)


def main():
    """
    Main entry point

    Usage:
        python fonetizer.py input.txt                    # CSV to stdout
        python fonetizer.py input.txt output.csv         # CSV to file
        python fonetizer.py input.txt output.xlsx        # Excel with formatting
        python fonetizer.py input.txt > output.csv       # CSV via redirection
    """
    # Read input
    if len(sys.argv) > 1:
        with open(sys.argv[1], 'r', encoding='utf-8') as f:
            lines = f.readlines()
    else:
        lines = sys.stdin.readlines()

    # Parse all phrases
    phrases = []
    for line in lines:
        line = line.strip()
        if not line:
            continue

        try:
            start, text = parse_input_line(line)
            syllables = process_phrase(text)
            phrases.append((start, syllables))
        except ValueError as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)

    # Determine output format
    if len(sys.argv) > 2:
        # Output filename specified as second argument
        output_file = sys.argv[2]
        if output_file.endswith('.xlsx'):
            # Generate formatted Excel file
            build_excel_table(phrases, output_file)
            print(f"Excel file created: {output_file}", file=sys.stderr)
        else:
            # Generate CSV file
            table = build_table(phrases)
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(table)
            print(f"CSV file created: {output_file}", file=sys.stderr)
    else:
        # Output to stdout (CSV format)
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        table = build_table(phrases)
        print(table)


if __name__ == '__main__':
    main()
